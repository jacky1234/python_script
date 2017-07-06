from openpyxl import Workbook
from openpyxl.compat import range

try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

wb = Workbook()
dest_filename = 'empty_book2.xlsx'
ws1 = wb.active  # 第一个表
ws1.title = "range names"  # 第一个表命名
# 遍历第一个表的1到40行，赋值一个600内的随机数
for row in range(1, 40):
    ws1.append(range(60))
ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}{1}".format(get_column_letter(col), row))
print(ws3['AA10'].value)

wb.save(filename=dest_filename)
